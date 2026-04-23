from fastapi import FastAPI, Request
import uvicorn  # сервер для запуска питон приложений
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, \
    StreamingResponse  # 1 - отправление html страниц, 2 - перенаправление юзера по разным страницам
from database import get_clients_page, get_total_count, get_contract_id, update_par, search_dog, get_dog_payments, \
    verify_windows_login, get_user_otd, add_dog_payment, delete_dog_payments, get_dog_payments1С, get_ds_data, \
    get_contract_files, get_user_connection, get_podr_list, get_user_id
from starlette.middleware.sessions import SessionMiddleware  # созданий сессий, запоминание что пользователь вошел
from starlette.middleware.base import BaseHTTPMiddleware  # проверка авторизации перед каждым запросом
import os
import shutil
from fastapi import UploadFile, File, Form
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font

os.environ.pop('http_proxy', None)
os.environ.pop('https_proxy', None)
os.environ.pop('HTTP_PROXY', None)
os.environ.pop('HTTPS_PROXY', None)
os.environ.pop('ftp_proxy', None)
os.environ.pop('FTP_PROXY', None)
os.environ.pop('all_proxy', None)
os.environ.pop('ALL_PROXY', None)
os.environ['NO_PROXY'] = '*'
os.environ['no_proxy'] = '*'

# создание веб-приложения
app = FastAPI()


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):  # функция выполняется для каждого запроса
        # список путей доступных без авторизации - в нашем случае это только страница входа, иначе бы перед каждым запросом кидал страницу входа
        public_paths = ["/login", "/logout", "/choose"]

        # если путь не в списке публичных, то проверяем есть ли пользователь в сессии, если да - то проходим, если нет - то отправка на логиниться
        if request.url.path not in public_paths:
            user = request.session.get("user")
            if not user:
                return RedirectResponse("/login", status_code=303)

        response = await call_next(request)
        return response


# добавляет класс работу каждый раз перед запросом
app.add_middleware(AuthMiddleware)

# для создания сессии юзера, для сохранения данных между запросами
app.add_middleware(
    SessionMiddleware,
    secret_key="arenkovaan",
    max_age=None,  # время работы сессии в секундах
    session_cookie="session",
    same_site="strict",
    https_only=False
)


# обработчик отправки логина и пароля
@app.post("/login")
async def login_post(request: Request):
    form = await request.form()  # забирает л/п
    username = form.get("username")  # достает логин
    password = form.get("password")  # достает пароль
    remote_host = "gazprosql"  # указывает на каком сервере проверять подключение

    if verify_windows_login(remote_host, username, password):  # проверяет по функции проверки из database л/п
        otd = get_user_otd(username)  # запоминает номер отдела у логина который ввели
        id_user = get_user_id(username)
        request.session["user"] = {"login": username, "otd": otd, "id_user": id_user}  # запомниает логин пользователя
        print(f"otd из функции: {otd}")
        return RedirectResponse("/choose", status_code=303)  # и отправляет чувака на главную с таблицу
    else:  # если не входит чувак то по новой ввод л/п
        return HTMLResponse(content="""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Вход в систему</title>
            <style>
                .login-body {
                    font-family: Arial;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100vh;
                    margin: 0;
                }
                .login-form {
                    border: 2px solid #1073b7;
                    padding: 30px;
                    border-radius: 10px;
                    width: 300px;
                }
                .login-input {
                    width: 100%;
                    padding: 8px;
                    margin: 10px 0;
                    box-sizing: border-box;
                }
                .login-button {
                    background: #1073b7;
                    color: white;
                    border: none;
                    padding: 10px;
                    width: 60%;
                    cursor: pointer;
                    margin-top: 10px;
                    border-radius: 4px;
                    display: block;
                    margin-left: auto;
                    margin-right: auto;
                }
                .login-error {
                    color: red;
                    margin-top: 10px;
                    text-align: center;
                }
                #seepassword {
                    position: absolute;
                    right: 10px;
                    top: 50%;
                    transform: translateY(-50%);
                    cursor: pointer;
                    font-size: 12px;
                    background: white;
                    padding: 0 5px;
                }
                .h2 {
                    text-align: center;
                    color: #0952a0;
                }
            </style>
        </head>

        <body class="login-body">
            <div class="login-form">
                <h2 class="h2">Вход в систему</h2>
                <form method="post" action="/login">
                    <input class="login-input" type="text" name="username" placeholder="Логин" required>
                    <div style="position: relative; width: 100%;">
                        <input class="login-input" type="password" name="password" id="password" placeholder="Пароль" required>
                        <span onclick="togglePassword()" id="seepassword">Показать</span>
                    </div>
                    <button class="login-button" type="submit">Войти</button>
                </form>
                <p class="login-error">Неверный логин или пароль</p>
            </div>

            <script>
            function togglePassword() {
                var passwordField = document.getElementById('password');
                passwordField.type = passwordField.type === 'password' ? 'text' : 'password';
            }
            </script>
        </body>
        </html>
        """)

# окно ввода логина и пароля
@app.get("/login", response_class=HTMLResponse)
def login_page():
    return """
    <html>
    <head>
        <title>Вход в систему</title>
        <style>
            .login-body {
                font-family: Arial;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
            }
            .login-form {      /* рамка */
                border: 2px solid #1073b7;
                padding: 30px;
                border-radius: 10px;
                width: 300px;
            }
            .login-input {
                width: 100%;
                padding: 8px;
                margin: 10px 0;
                box-sizing: border-box;
            }
            .login-button {
                background: #1073b7;
                color: white;
                border: none;
                padding: 10px;
                width: 60%;
                cursor: pointer;
                margin-top: 10px;
                border-radius: 4px;
                display: block;
                margin-left: auto;
                margin-right: auto;
            }
            #seepassword {
                position: absolute;
                right: 10px;
                top: 50%;
                transform: translateY(-50%);
                cursor: pointer;
                font-size: 12px;
                background: white;
                padding: 0 5px;
            }
            .h2 {
                text-align: center;
                color: #0952a0;
            }
        </style>
    </head>

    <body class="login-body">
    <div class="login-form">
        <h2 class="h2">Вход в систему</h2>
        <form method="post" action="/login">
            <input class="login-input" type="text" name="username" placeholder="Логин" required>
            <div style="position: relative;">
                <input class="login-input" type="password" name="password" id="password" placeholder="Пароль" required>
                <span onclick="togglePassword()" id="seepassword">Показать</span>
            </div>
            <button class="login-button" type="submit">Войти</button>
        </form>
    </div>

    <!-- функция для нажатия кнопки показать и пароль станет видимым. если при нажатии кнопки звездочки, то будут буквы и наоборот -->
    <script>
    function togglePassword() {
        var passwordField = document.getElementById('password');
        if (passwordField.type === 'password') {
            passwordField.type = 'text';
        } 
        else {
            passwordField.type = 'password';
        }
    }
    </script>"""

@app.get("/logout")
async def logout(request: Request):
    # очищаем сессию
    request.session.clear()
    # отправляем на страницу входа
    return RedirectResponse("/login", status_code=303)


# функция главного экрана с пагинацией
@app.get("/", response_class=HTMLResponse)
def home(request: Request, page: int = 1, page_size: int = 4000):
    contracts = get_clients_page(request, page, page_size)  # занесение списка из строк которые будем выводить
    total_count = get_total_count(request)  # занесение общего числа строк

    # расссчитывает сколько всего должно быть страниц на всё количество с округлением вверх
    total_pages = (total_count + page_size - 1) // page_size
    podr_list = get_podr_list(request)
    podr_options = '<option value="">Все подразделения</option>'
    for podr in podr_list:
        podr_options += f'<option value="{podr["id"]}">{podr["name"]}</option>'

    html = """  
    <html>
    <head><title>Договора в ЕИС</title>
    <style>
    .table-columns {
        width: 100%;     /* ширина таблицы */
        max-width: 100%;    /* максимально занимаемая ширина таблицы */
        table-layout: fixed;     /* фиксированные размеры столбцов */
        border-collapse: collapse;    /* делает двойные границы одной линией */
    }

    .table-columns th:nth-child(1),
    .table-columns td:nth-child(1) { width: 5%; }
    .table-columns th:nth-child(2),
    .table-columns td:nth-child(2) { width: 15%; }
    .table-columns th:nth-child(3),
    .table-columns td:nth-child(3) { width: 10%; }
    .table-columns th:nth-child(4),
    .table-columns td:nth-child(4) { width: 10%; }
    .table-columns th:nth-child(5),
    .table-columns td:nth-child(5) { width: 25%; }
    .table-columns th:nth-child(6),
    .table-columns td:nth-child(6) { width: 35%; }

    th, td {    /* th - заголовки свойства самих ячеек */
        white-space: normal;
        word-wrap: break-word;
        overflow-wrap: break-word;
    }
    .h1 {
        background: #e4f0fb; /* Цвет фона под заголовком */
        color: #0952a0; /* Цвет текста */
        padding: 8px; /* Поля вокруг текста */
    }
    .column-names {
        color: #0952a0; /* Цвет текста */
    }
    tr:hover {
        background-color: #f5f5f5; /* цвет при наведении */
        cursor: pointer;           /* курсор в виде руки */
    }
    .pagination {      /* стиль для одной иконки страницы */
        margin: 20px 0;
        text-align: center;
    }  
    .pagination a {        /* стиль для всех иконок страниц */
        display: inline-block;
        padding: 5px 10px;
        margin: 0 2px;
        border: 1px solid #ddd;
        text-decoration: none;
    }
    .pagination a.active {     /* стиль для активной страницы, которая нажата */
        background: #032c57;
        color: white;
    }
    .button-searchdog {
        padding: 10px 20px;
        background-color: #1073b7;
        color: white;
        border: none;
        cursor: pointer;
        font-size: 16px;
        border-radius: 4px;
    }
    .button-in-window {
        background-color: #1073b7;
        color: white;
        border: none;
        cursor: pointer;
        font-size: 14 px;
        padding: 5px 10px;
        border-radius: 4px;
    }
    .button-exit {      /*кнопка выхода из ученой записи*/
        background-color: #1073b7;
        color: white;
        border: none;
        cursor: pointer;
        font-size: 16px;
        padding: 10px 20px;
        text-decoration: none;
        border-radius: 4px;
    }
    .button-switch {
        background-color: #28a745;
        color: white;
        border: none;
        cursor: pointer;
        font-size: 16px;
        padding: 8px 16px;
        border-radius: 4px;
        text-decoration: none;
        transition: background 0.2s;
    }
    .button-switch:hover {
        background-color: #218838;
    }
    </style>
    </head>
    """
    html += f"""
    <body>
        <div style="background: #e4f0fb; padding: 8px 15px; display: flex; justify-content: space-between; align-items: center;">
            <h1 style="margin: 0; color: #0952a0; font-size: 28px;">Список договоров подлежащих публикации в ЕИС</h1>
            <div style="display: flex; gap: 12px;">
                <button class="button-searchdog" onclick="opendialogwindow()">Найти договор</button>
                <button class="button-switch" onclick="goToChoose()">Сменить программу</button>
                <button class="button-exit" onclick="logout()">Выйти</button>
            </div>
        </div>
        
        <div style="font-size: 15px; color: #666; margin-bottom: 5px;">
            Всего договоров: <strong>{total_count}</strong>
        </div>
        <!--ДИАЛОГОВОЕ ОКНО ПОИСКА ДОГОВОРА------------------>
        <dialog id="dialogwindow" style="widht: 25%; border: 2px solid black;">
            <p style="color: #1073b7; font-size: 20px; margin: 10px 0px;"><strong>Поиск договора</strong></p>
            <div style="margin-bottom: 10px;">
                <label>№ договора(числовой)</label>
                <input type="text" id="numberdog" placeholder="Введите номер договора">
            </div>
            <div style="margin-bottom: 10px;">
                <label>№ контрагента</label>
                <input type="text" id="numberkontr" placeholder="Введите номер контрагента">
            </div>
            <div style="display: flex; gap: 10px; margin-bottom: 10px;">
                <label>Дата договора</label>
                <div style="display: flex; gap: 10px;">
                    <input type="date" id="date_from" placeholder="с">
                    <span>-</span>
                    <input type="date" id="date_to" placeholder="по">
                </div>
            </div>
            <div style="display: flex; gap: 7px; margin-top: 5px;">
                <input type="checkbox" id="publ">
                <label>Только подлежащие публикации</label>
            </div>
            <div style="margin-bottom: 10px; display: flex; gap: 10px;">
                <label>Привязка файлов</label>
                <div style="display: flex; gap: 10px;">
                    <input type="date" id="file_date_from" placeholder="с">
                    <span>-</span>
                    <input type="date" id="file_date_to" placeholder="по">
                </div>
            </div>
            <div style="margin-bottom: 10px; display: flex; gap: 10px;">
                <label>Сумма договора</label>
                <div style="display: flex; gap: 10px;">
                    <input type="text" id="sum_from" placeholder="от" size="10">
                    <span>-</span>
                    <input type="text" id="sum_to" placeholder="до" size="10">
                </div>
            </div>
            <div style="display: flex; gap: 7px; margin-top: 5px;">
                <label>Подразделение</label>
                <select id="podr_select" style="width: 100%; padding: 5px;">
                    {podr_options}
                </select>
            </div>
            <div style="margin-bottom: 10px;">
                <label>Предмет договора</label>
                <input type="text" id="pr_dog" placeholder="Введите текст" style="width: 60%; padding: 5px;">
            </div>
            <div style="display: flex; gap: 7px; margin-top: 5px;">
                <input type="checkbox" id="gazsrv">
                <label>Нижегородоблгаз Сервис</label>
            </div>
            <div style="display: flex; gap: 7px; margin-top: 5px;">
                <input type="checkbox" id="search_archive">
                <label>Поиск в архиве</label>
            </div>
            <div style="gap: 20px;">
                <button class="button-in-window" onclick="searchdog()">Найти</button>
                <button class="button-in-window" onclick="goback()">Отмена</button>
            </div>
        </dialog>
        <!--ЗАКРЫТИЕ ДИАЛОГОВОГО ОКНА ПОИСКА ДОГОВОРА------------------>

        <table border="1" class="table-columns">
    """
    html += '<table border="1" class="table-columns">'
    if contracts:
        # Заголовки
        html += "<tr style='background-color: #f2f2f2;'>"
        html += "<th style='padding: 10px; border: 1px solid;'>ID</th>"  # ← добавляем вручную
        for key in contracts[0].keys():
            if key != 'ID договора':
                html += f"<th style='padding: 10px; border: 1px solid;'>{key}</th>"
        html += "</tr>"

        # Данные
        for contract in contracts:
            contract_id = contract['ID договора']  # для ссылки
            display_id = f"8{str(contract_id).zfill(6)}"

            html += f"<tr onclick='showContract({contract_id}, {page})'>"
            html += f"<td style='padding: 5px; border: 1px solid #ddd;'>{display_id}</td>"  # красивый ID
            for key, value in contract.items():
                if key != 'ID договора':
                    html += f"<td style='padding: 5px; border: 1px solid #ddd;'>{value}</td>"
            html += "<tr>"
    html += """</table> 

    <script>
    // функция для перехода на страницу договора при двойном клике по списку 
    function showContract(id, page) {
        window.location.href = '/contract/' + id + '?from_page=' + page;
    }

    // функция откытия диалогового окна после того как нажали на кнопку найти договор 
    // document - главная страница, getElementById('dialogwindow') - найти элемент по id, showModal() - открыть поверх станицы, затемняя фон  
    function opendialogwindow() {
        document.getElementById('dialogwindow').showModal()      
    }

    // функция закрытия окна через кнопку отмена, close() - закрыть окно 
    function goback() {
        document.getElementById('dialogwindow').close();
    }

    // const - переменная в дальнейшем не меняется, пишем когда не объявляли переменную ранее, value - взять то что написано в элементе
    function searchdog() {
        const numberdog = document.getElementById('numberdog').value;
        const numberkontr = document.getElementById('numberkontr').value;
        const dateFrom = document.getElementById('date_from').value;
        const dateTo = document.getElementById('date_to').value;
        const publ = document.getElementById('publ').checked ? '1' : '';
        const sumFrom = document.getElementById('sum_from').value;
        const sumTo = document.getElementById('sum_to').value;
        const podr = document.getElementById('podr_select').value;
        const prDog = document.getElementById('pr_dog').value;
        const gazsrv = document.getElementById('gazsrv').checked ? '1' : ''; 
        const searchArchive = document.getElementById('search_archive').checked ? '1' : '';

        let url = '/search?';
        let params = [];

        if (numberdog) params.push('numberdog=' + encodeURIComponent(numberdog));
        if (numberkontr) params.push('numberkontr=' + encodeURIComponent(numberkontr));
        if (dateFrom) params.push('date_from=' + encodeURIComponent(dateFrom));
        if (dateTo) params.push('date_to=' + encodeURIComponent(dateTo));
        if (publ) params.push('publ=' + publ);  
        if (sumFrom) params.push('sum_from=' + encodeURIComponent(sumFrom));  
        if (sumTo) params.push('sum_to=' + encodeURIComponent(sumTo));
        if (podr) params.push('podr=' + encodeURIComponent(podr)); 
        if (prDog) params.push('pr_dog=' + encodeURIComponent(prDog)); 
        if (gazsrv) params.push('gazsrv=' + gazsrv);
        if (searchArchive) params.push('search_archive=' + searchArchive);

        url += params.join('&');
        window.location.href = url;
        document.getElementById('dialogwindow').close();
    }

    function logout() {
        window.location.href = '/logout';
    }
    
    function goToChoose() {
        window.location.href = '/choose';
    }
    </script>

    <div class="pagination">
    """
    for i in range(1,
                   total_pages + 1):  # рассчитывает сколько кнопок пагинации надо создать. если total_pages = 4, то кнопок 4
        active_class = "active" if i == page else ""  # определяет активную кнопку. если i равняется номеру открытой страницы, то применяется к ней стиль active_class
        html += f'<a href="/?page={i}&page_size={page_size}" class="{active_class}">{i}</a>'  # создание самой кнопки страницы

    html += """
    </div>
    </body>
    </html>
    """
    return html


# функция перехода на старицу договора когда нашли 1 договор через кнопку найти договор
@app.get("/search", response_class=HTMLResponse)
def search_page(request: Request,
                numberdog: str = "",
                numberkontr: str = "",
                date_from: str = "",
                date_to: str = "",
                publ: str = "",
                sum_from: str = "",
                sum_to: str = "",
                podr: str = "",
                pr_dog: str = "",
                gazsrv: str = "",
                search_archive: str = ""
                ):
    results = search_dog(request,
                         numberdog,
                         numberkontr,
                         date_from,
                         date_to,
                         publ,
                         sum_from,
                         sum_to,
                         podr,
                         pr_dog,
                         gazsrv,
                         search_archive
                         )

    # Если нашли ровно один договор - сразу переходим на него
    if len(results) == 1:
        contract_id = results[0]['ID договора']
        # Не передаем параметры поиска, только from_page=search
        return HTMLResponse(f"""
            <script>
                window.location.href = '/contract/{contract_id}?from_page=search'
            </script>
            """)

    # сoбираем параметры для перехода (для таблицы результатов)
    params = []
    if numberdog:
        params.append(f'numberdog={numberdog}')
    if numberkontr:
        params.append(f'numberkontr={numberkontr}')
    if date_from:
        params.append(f'date_from={date_from}')
    if date_to:
        params.append(f'date_to={date_to}')
    if publ:
        params.append(f'publ={publ}')
    if sum_from:
        params.append(f'sum_from={sum_from}')
    if sum_to:
        params.append(f'sum_to={sum_to}')
    if podr:
        params.append(f'podr={podr}')
    if pr_dog:
        params.append(f'pr_dog={pr_dog}')
    if gazsrv:
        params.append(f'gazsrv={gazsrv}')
    if search_archive:
        params.append(f'search_archive={search_archive}')

    query_string = '&'.join(params)
    if query_string:
        query_string = '?' + query_string
    else:
        query_string = ''

    # Формируем HTML страницы
    html = f"""
    <html>
    <head><title>Результаты поиска</title>
    <style>
        .h1 {{
            background: #e4f0fb;
            color: #0952a0;
            padding: 8px;
        }}
        .button-back {{
            background: #1073b7;
            font-size: 18px;
            padding: 8px 20px;
            color: white;
            text-decoration: none;
            border: none;
            border-radius: 4px;
            display: inline-block;
            font-family: inherit;
        }}
        .button-back:hover {{
            background: #0952a0;
        }}
        .table-wrapper {{
            margin: 20px 0;
        }}
    </style>
    </head>
    <body>
    <div>
        <h1 class="h1">Результаты поиска</h1>
        <div style="display: flex; justify-content: space-between; margin-top: 20px; margin-bottom: 20px; text-align: left; gap">
            <a href="/" class="button-back">Назад к списку</a>
            <button onclick="exportToExcel()" class="button-back">Сформировать отчет</button>
        </div>
        <div style="margin-block: 3px;"><strong>Договор:</strong> {numberdog if numberdog else 'не указан'}</div>
        <div style="margin-bottom: 3px;"><strong>Контрагент:</strong> {numberkontr if numberkontr else 'не указан'}</div>
        <div style="margin-bottom: 3px;"><strong>Дата с:</strong> {date_from if date_from else 'не указана'}</div>
        <div style="margin-bottom: 3px;"><strong>Дата по:</strong> {date_to if date_to else 'не указана'}</div>
        <div style="margin-bottom: 3px;"><strong>Только подлежащие публикации:</strong> {'да' if publ else 'нет'}</div>
        <div style="margin-bottom: 3px;"><strong>Сумма от:</strong> {sum_from if sum_from else 'не указана'}</div>
        <div style="margin-bottom: 3px;"><strong>Сумма до:</strong> {sum_to if sum_to else 'не указана'}</div>
        <div style="margin-bottom: 3px;"><strong>Подразделение:</strong> {podr if podr else 'не указано'}</div>
        <div style="margin-bottom: 3px;"><strong>Предмет договора:</strong> {pr_dog if pr_dog else 'не указан'}</div>
        <div style="margin-bottom: 3px;"><strong>Нижегородоблгаз Сервис:</strong> {'да' if gazsrv else 'нет'}</div>
        <div style="margin-bottom: 3px;"><strong>Поиск в архиве:</strong> {'да' if search_archive else 'нет'}</div>
        <div style="font-size: 15px; color: #666;">
            Найдено договоров: <strong>{len(results)}</strong>
        </div>
        <div class="table-wrapper">
            <table border="1" style="border-collapse: collapse; width: 100%;">
    """

    if results:
        # Заголовки
        html += "<tr style='background-color: #f2f2f2'>"
        html += "<th style='padding: 10px; border: 1px solid;'>ID</th>"
        for key in results[0].keys():
            if key != 'ID договора':
                html += f"<th style='padding: 10px; border: 1px solid;'>{key}</th>"
        html += "</tr>"

        # Данные таблицы
        clean_params = query_string[1:] if query_string.startswith('?') else query_string

        for contract in results:
            contract_id = contract['ID договора']
            display_id = f"8{str(contract_id).zfill(6)}"

            if clean_params:
                html += f"""<tr ondblclick="window.location.href='/contract/{contract_id}?from_page=search&{clean_params}'">"""
            else:
                html += f"""<tr ondblclick="window.location.href='/contract/{contract_id}?from_page=search'">"""

            # Выводим ячейки
            html += f"<td style='padding: 5px; border: 1px solid #ddd;'>{display_id}</td>"
            for key, value in contract.items():
                if key != 'ID договора':
                    html += f"<td style='padding: 5px; border: 1px solid #ddd;'>{value}</td>"
            html += "</tr>"
    else:
        html += f"<tr><td colspan='6' style='padding: 20px; text-align: center;'>Договоры не найдены</td></tr>"
    html += f"""
            </table>
        </div>
    </div>

    <script>
        function exportToExcel() {{
            const urlParams = new URLSearchParams(window.location.search);
            fetch(`/api/contract/export?${{urlParams.toString()}}`)
                .then(response => response.blob())
                .then(blob => {{
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `реестр_договоров.xlsx`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                }})
                .catch(error => {{
                    alert('Ошибка при формировании отчета: ' + error);
                }});
        }}
    </script>
    </body>
    </html>
    """
    return html


# функция для показа информации клиента
@app.get("/contract/{contract_id}", response_class=HTMLResponse)
def contract_page(request: Request, contract_id: int, from_page: str = "1"):
    # присваивание результата функции в переменную
    contract = get_contract_id(request, contract_id)
    if not contract:
        return "<h1>Договор не найден</h1>"

    # Заменяем все None на пустые строки
    for key in contract:
        if contract[key] is None:
            contract[key] = ''

    # Получаем параметры из URL
    from_page_param = request.query_params.get('from_page', '1')
    search_numberdog = request.query_params.get('numberdog', '')
    search_numberkontr = request.query_params.get('numberkontr', '')
    search_date_from = request.query_params.get('date_from', '')
    search_date_to = request.query_params.get('date_to', '')
    search_sum_from = request.query_params.get('sum_from', '')
    search_sum_to = request.query_params.get('sum_to', '')
    search_publ = request.query_params.get('publ', '')
    search_privyazka = request.query_params.get('privyazka', '')
    search_podr = request.query_params.get('podr', '')
    search_pr_dog = request.query_params.get('pr_dog', '')
    search_service = request.query_params.get('service', '')
    search_archive = request.query_params.get('search_archive', '')
    search_file_date_from = request.query_params.get('file_date_from', '')
    search_file_date_to = request.query_params.get('file_date_to', '')

    # Собираем параметры для back_url
    params = []
    if search_numberdog:
        params.append(f'numberdog={search_numberdog}')
    if search_numberkontr:
        params.append(f'numberkontr={search_numberkontr}')
    if search_date_from:
        params.append(f'date_from={search_date_from}')
    if search_date_to:
        params.append(f'date_to={search_date_to}')
    if search_sum_from:
        params.append(f'sum_from={search_sum_from}')
    if search_sum_to:
        params.append(f'sum_to={search_sum_to}')
    if search_publ:
        params.append(f'publ={search_publ}')
    if search_privyazka:
        params.append(f'privyazka={search_privyazka}')
    if search_podr:
        params.append(f'podr={search_podr}')
    if search_pr_dog:
        params.append(f'pr_dog={search_pr_dog}')
    if search_service:
        params.append(f'service={search_service}')
    if search_archive:
        params.append(f'search_archive={search_archive}')
    if search_file_date_from:
        params.append(f'file_date_from={search_file_date_from}')
    if search_file_date_to:
        params.append(f'file_date_to={search_file_date_to}')

    has_search_params = len(params) > 0

    # Формируем back_url
    if from_page_param.startswith('search'):
        if has_search_params:
            back_url = '/search?' + '&'.join(params)
        else:
            back_url = '/'
    else:
        back_url = f'/?page={from_page_param}'

    ds_data = get_ds_data(request, contract_id)
    payments = get_dog_payments(request, contract_id)
    payments1C = get_dog_payments1С(request, contract_id)
    total_sum_1c = 0  # подсчет общей суммы
    if payments1C:
        for payment in payments1C:
            total_sum_1c += float(payment.get('s_opl', 0))

    konk_value = contract.get('Конкурс', 0)  # берем значение из столбца Конкурс, если значения нет - 0
    is_checked = 'checked' if konk_value == 1 else ''

    num_z_value = contract.get('№ закупки ЕИС', '')
    numzalel_value = contract.get('№ закупю на эл.пл.', '')

    przak_value = contract.get('Прямая закупка', 0)
    przak_checked = 'checked' if przak_value == 1 else ''

    beznds_value = contract.get('Без НДС', 0)
    beznds_checked = 'checked' if beznds_value == 1 else ''

    opl_value = contract.get('Оплачено', 0)
    opl_checked = 'checked' if opl_value == 1 else ''

    eis_value = contract.get('Публикация в ЕИС', 0)
    eis_checked = 'checked' if eis_value == 1 else ''

    osn_value = contract.get('Основание', '')
    gpz_value = contract.get('ГПЗ', '')
    uid_value = contract.get('UID', '')
    ppz_value = contract.get('ППЗ', '')
    # Преобразуем код stz в название
    stz_value = contract.get('stz', '0')
    stz_dict = {
        '0': 'не указан',
        '1': 'БДР',
        '2': 'Инвестиции',
        '3': 'Прочее',
        '4': 'В рамках агентского'
    }
    stz_name = stz_dict.get(str(stz_value), 'не указан')
    publ_value = contract.get('Публикация', 0)
    publ_checked = 'checked' if publ_value == 1 else ''
    publ_d_value = contract.get('Дата публикации', '')

    prol_value = contract.get('Пролонгация', 0)
    prol_checked = 'checked' if prol_value == 1 else ''

    d_end_raw = contract.get('Дата заверш. договора', '')
    d_end_value = d_end_raw.strftime('%Y-%m-%d') if d_end_raw else ''

    sposobzak_value = contract.get('Способ закупки', '')
    VIdZAK_value = contract.get('Вид закупки', '')
    numzak_value = contract.get('№ КЗ', '')
    predlog_value = contract.get('Формат закупки', '')
    dat_docosznak = contract.get('Дата дог.осн.закупки', '')
    dat_docosznak_value = dat_docosznak.strftime('%Y-%m-%d') if dat_docosznak else ''
    smsp_value = contract.get('СМСП', '')
    num_docosnzak_value = contract.get('№ дог.осн.закупки', '')
    OSTNEKONZAK_value = contract.get('Код основания некн.зак.', '')
    okpd2_value = contract.get('ОКПД2', '')
    subectzak_value = contract.get('SUBECTZAK', '')

    s_dog_okz_value = contract.get('Сумма договора ОКЗ', '')
    s_ds_value = contract.get('Сумма с ДС', '')
    date_izv_value = contract.get('Дата извещения', '')

    agent_value = contract.get('Агентский договор', 0)
    agent_checked = 'checked' if agent_value == 1 else ''

    smsp_okz_value = contract.get('Загрузка среди СМСП', 0)
    smsp_okz_checked = 'checked' if smsp_okz_value == 1 else ''

    sysnum_value = contract.get('Системный номер', '')
    d_work_value = contract.get('Рабочая дата', '')

    predlog_txt_value = contract.get('№ предложения', '')
    display_id = f"8{str(contract_id).zfill(7)}"
    # оформление страницы договора
    html = f"""
    <html>
    <head><title>Договор {contract_id}</title>
    <style>
    .button-back-page {{
        background: #1073b7; /* Цвет фона */
        font-size: 18px; /* Размер текста */
        padding: 5px 20px; /* Поля вокруг текста */
        color: white;
        text-decoration: none;
        border-radius: 4px; 
    }}
    .button-back-page:hover {{
        background: #0952a0;
        border-radius: 4px; 
    }}
    .h1 {{
        background: #e4f0fb; /* Цвет фона под заголовком */
        color: #0952a0; /* Цвет текста */
        padding: 8px; /* Поля вокруг текста */
    }}
    .info-container {{    /* фон под информацией */
        background: #f8f9fa;  /* Светло-серый фон */
        padding-top: 10px;
        padding-left: 20px;
        padding-right: 20px;
        padding-bottom: 10px; 
        border-radius: 10px;
        margin: 20px 0;
        display: flow-root; 
    }}
    .konk-Сheckbox {{   /* квадратик с галочкой */
        width: 24px;
        height: 24px;
        cursor: pointer;
    }}
    .button-save {{
        background: #1073b7;       /* фон кнопки */ 
        font-size: 18px;      /* размер текста */
        padding: 5px 20px;      /* отступы для текста внутри кнопки */
        color: white;    /* цвет текста */
        text-decoration: none;      /* убирает подчерквание */
        cursor: pointer;     /* курсор руки при наведении */
        border-radius: 4px;     /* сруление углов кнопки */
        font-family: inherit;    /* наследование */
        border: 1px solid #1073b7;
    }}
    .button-save:hover {{
        background: #0952a0;    /* темно синий при наведении мыши */
        border: 1px solid #1073b7
    }}
    .button-adddeletepayment {{
        width: 30px; 
        height: 30px; 
        background: #1073b7; 
        color: white; 
        border: none; 
        border-radius: 3px; 
        cursor: pointer; 
        font-size: 18px;
    }}
    .button-docs {{
        background: #1073b7;
        color: white;
        border: none;
        padding: 8px 8px;
        border-radius: 4px;
        cursor: pointer;
        font-size: 18px;
        margin-top: 10px;
    }}
    .button-docs:hover {{
        background: #0952a0;
    }}

    </style>
    </head>

    <body>
    <h1 class="h1">Информация по договору {display_id}</h1>

    <div class="info-container">

        <div style="display: grid; grid-template-columns: 1fr 1fr 0.3fr; gap: 20px; margin: 0px; width: 100%">   <!--начало двух столбцов-->
        <div>    <!--контейнер левого столбца-->
        <!-- СВЕРХУ СЛЕВА -------------------------->
        <div style="line-height: 24px; width: 100%;">
            <div style="width: 100%; display: flex; gap: 90px;">
                <span><strong>№ договора:</strong> {contract.get('№ договора', 'Нет данных')}</span>
                <span><strong>№ контрагента:</strong> {contract.get('№ контрагента', 'Нет данных')}</span>
            </div>

            <div style="display: flex; gap: 50px; align-items: center; width: 100%">
                <div><strong>Дата регистрации:</strong> {str(contract.get('Дата регистрации', ''))[:10]}</div>
                <div><strong>Дата договора:</strong> {str(contract.get('Дата договора', ''))[:10]}</div>
                <div style="display: flex; align-items: center; gap: 10px;">
                    <input type="checkbox" id="konkCheckbox" {is_checked}>
                    <label for="konkCheckbox" style="cursor: pointer;"><strong>Конкурс</strong></label>
                </div>
            </div>

            <div style="width: 100%"><strong>Подразделение:</strong> {contract.get('Подразделение', 'Нет данных')}</div>
            <div style="width: 100%; word-wrap: break-word;"><strong>Предмет договора:</strong> {contract.get('Предмет договора', 'Нет данных')}</div>
            <div style="display: flex; gap: 40px; width: 100%">
                <span><strong>Дата начала:</strong> {str(contract.get('Дата начала', ''))[:10]}</span>
                <span><strong>Дата конца:</strong> {str(contract.get('Дата конца', ''))[:10]}</span>
                <span><strong>Сумма договора:</strong> {contract.get('Сумма договора', 'Нет данных')} рублей</span>
            </div>
        </div>
        <!--КОНЕЦ СВЕРХУ СЛЕВА--------------------->

        <!-- КОНТРАГЕНТ ----------------------->
        <div style="border: 2px solid #1073b7; padding: 15px; border-radius: 8px; margin: 15px 0 10px 0;">
            <h3 style="margin: 0 0 10px 0; color: #0952a0;">Контрагент</h3>

            <div style="line-height: 25px;">
                <div><strong>Наименование:</strong> {contract.get('Наименование', 'Нет данных')}</div>

                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin: 0px;">

                        <div style="line-height: 25px;">
                            <div><strong>ИНН:</strong> {contract.get('ИНН', '')}</div>
                            <div><strong>Расч.счет:</strong> {contract.get('Расч.счет', '')}</div>
                        </div>

                        <div style="line-height: 25px;">
                            <div><strong>КПП:</strong> {contract.get('КПП', '')}</div>
                            <div><strong>Телефон/факс:</strong> {contract.get('Телефон/факс', '')}</div>
                        </div>
                </div>

                <div><strong>БИК:</strong> {contract.get('БИК', 'Нет данных')}</div>
            </div>
        </div>
        <!--КОНЕЦ КОНТРАГЕНТА--------------------->

        <!-- НИЖЕ КОНТРАГЕНТА ---------------->
        <div style="line-height: 28px;">
            <div style="display: flex; gap: 25px;">
                <div><label><strong>№ закупки ЕИС:</strong></label>
                <input type="text" id="num_z" value="{num_z_value}" size="15"></div>

                <div><label><strong>№ закуп. на эл.пл.:</strong></label>
                <input type="text" id="numzalel" value="{numzalel_value}" size="15"></div>
            </div>

            <div style="display: flex; gap: 25px;">
                <div><input type="checkbox" id="przakcheckbox" {przak_checked}>
                <label for="przakcheckbox" style="cursor: pointer;"><strong>Прямая закупка:</strong></label></div>

                <div><label><strong>Основание:</strong></label>
                <input type="text" id="osn" value="{osn_value}" size="15"></div>

                <div><label><strong>ГПЗ:</strong></label>
                <input type="text" id="gpz" value="{gpz_value}" size="15"></div>
            </div>

            <div style="display: flex; gap: 25px;">
                <div><label><strong>UID:</strong></label>
                <input type="text" id="uid" value="{uid_value}" size="20"></div>

                <div><label><strong>ППЗ:</strong></label>
                <input type="text" id="ppz" value="{ppz_value}" size="15"></div>
            </div>

            <div style="display: flex; gap: 25px;">
                <div style="display: flex; gap: 5px; align-items: center; margin-bottom: 5px; size="20"">
                    <label for="stz_value"><strong>Статьи затрат:</strong></label>
                    <input type="text" id="stz_value" value="{stz_name}" readonly style=" cursor: default;">
                </div>

                <div style="display: flex; gap: 30px; align-items: center; margin-bottom: 5px;">
                    <div>
                        <input type="checkbox" id="publ_checkbox" {publ_checked}>
                        <label for="publ_checkbox"><strong>Опубликовано</strong></label>
                    </div>
                    <div id="publ_date_div" style="display: {'block' if publ_value == 1 else 'none'};">
                        <label><strong>Дата публикации</strong></label>
                        <input type="date" id="publ_date" value="{publ_d_value}">
                    </div>
                </div>
            </div>
        </div>
        <!--КОНЕЦ НИЖЕ КОНТРАГЕНТА------------------>
        </div>     <!-- закрытие левого столбца-->

        <script>
        // Показать/скрыть поле даты публикации
        const publCheckbox = document.getElementById('publ_checkbox');
        const publDateDiv = document.getElementById('publ_date_div');

        if (publCheckbox) {{
            publCheckbox.addEventListener('change', function() {{
                if (publDateDiv) {{
                    publDateDiv.style.display = this.checked ? 'block' : 'none';
                }}
            }});
        }}
        </script>

        <div>       <!-- открытие правого столбца -->
        <!-- СВЕРХУ СПРАВА-----------------> 
        <div style="">
            <div style="display: flex; gap: 30px; margin-bottom: 10px;">
                <div><input type="checkbox" id="prol" {prol_checked}>
                <label for="prol" style="cursor: pointer;"><strong>Пролонгация:</strong></label></div>

                <div><input type="checkbox" id="beznds" {beznds_checked}>
                <label for="beznds" style="cursor: pointer;"><strong>Без НДС</strong></label></div>

                <div><strong>Код ОБД НСИ:</strong> {contract.get('Код ОБД НСИ', 'Нет данных')}</div>
            </div>

            <div style=" display: flex; flex-direction: column;">
                <div><input type="checkbox" id="opl" {opl_checked}>
                <label for="opl" style="cursor: pointer;"><strong>Оплачено</strong></label></div>

                <div style="margin-bottom: 10px;"><input type="checkbox" id="eis" {eis_checked}>
                <label for="eis" style="cursor: pointer; "><strong>Публикация в ЕИС</strong></label></div>
            </div>

            <div style="margin-bottom: 10px;">
                <label><strong>Дата заверш. договора:</strong></label>
                <input type="date" id="d_end" value="{d_end_value}">

                <div>
                    <label for="statusD"><strong>СТАТУС</strong></label>
                    <select name="statusD" id="status">
                    <option value="2">Подписан</option>
                    <option value="4">Закрыт</option>
                    </select>
                </div>
            </div>
        </div>
        <!--КОНЕЦ СВЕРХУ СПРАВА----------------->

        <!-- РЕКВИЗИТЫ------------------->
        <div style="width: 90%; border: 2px solid #1073b7; padding: 15px 15px 5px 20px; border-radius: 8px; margin: 10px 0;">
            <h3 style="margin: 0 0 10px 0; color: #0952a0;">Реквизиты</h3>

            <div style="line-height: 28px;">
                <div style="display: flex; gap: 50px;">
                    <div><form action="#" style="margin-bottom: 0;">
                        <label for="sposobzak"><strong>Способ закупки</strong></label>
                        <select name="sposobzak" id="sposobzak">
                        <option value="">   </option>
                        <option value="001" {'selected' if sposobzak_value == '001' else ''}>001</option>
                        <option value="002" {'selected' if sposobzak_value == '002' else ''}>002</option>
                        <option value="003" {'selected' if sposobzak_value == '003' else ''}>003</option>
                        <option value="004" {'selected' if sposobzak_value == '004' else ''}>004</option>
                        <option value="005" {'selected' if sposobzak_value == '005' else ''}>005</option>
                        <option value="006" {'selected' if sposobzak_value == '006' else ''}>006</option>
                        <option value="007" {'selected' if sposobzak_value == '007' else ''}>007</option>
                        <option value="008" {'selected' if sposobzak_value == '008' else ''}>008</option>
                        <option value="009" {'selected' if sposobzak_value == '009' else ''}>009</option>
                        <option value="010" {'selected' if sposobzak_value == '010' else ''}>010</option>
                        <option value="011" {'selected' if sposobzak_value == '011' else ''}>011</option>
                        <option value="012" {'selected' if sposobzak_value == '012' else ''}>012</option>
                        <option value="013" {'selected' if sposobzak_value == '013' else ''}>013</option>
                        <option value="014" {'selected' if sposobzak_value == '014' else ''}>014</option>
                        <option value="015" {'selected' if sposobzak_value == '015' else ''}>015</option>
                        <option value="016" {'selected' if sposobzak_value == '016' else ''}>016</option>
                        <option value="017" {'selected' if sposobzak_value == '017' else ''}>017</option>
                        <option value="018" {'selected' if sposobzak_value == '018' else ''}>018</option>
                        </select>
                    </form> 
                    </div>

                    <div >
                        <form action="#" style="margin-bottom: 0;">
                            <label for="VIdZAK"><strong>Вид закупки</strong></label>
                            <select name="VIdZAK" id="VIdZAK">
                                <option value="">   </option>
                                <option value="001" {'selected' if VIdZAK_value == '001' else ''}>001</option>
                                <option value="002" {'selected' if VIdZAK_value == '002' else ''}>002</option>
                            </select>
                        </form> 
                    </div>
                </div>

                <div style="display: flex; gap: 50px;">
                    <div style="margin-bottom: 0;">
                        <label><strong>№ КЗ:</strong></label>
                        <input type="text" id="numzak" value="{numzak_value}" size="15">
                    </div>

                    <div>
                        <form action="#" style="margin-bottom: 0;">
                            <label for="predlog"><strong>Формат закупки</strong></label>
                        <select name="predlog" id="predlog">
                            <option value="">   </option>
                            <option value="opened" {'selected' if predlog_value == 'opened' else ''}>Открытая</option>
                            <option value="closed" {'selected' if predlog_value == 'closed' else ''}>Закрытая</option>
                        </select>
                        </form>
                    </div>
                </div>

                <div style="display: flex; gap: 50px;">
                    <div>
                        <label><strong>Дата док.осн.зак.:</strong></label>
                        <input type="date" id="dat_docosznak" value="{dat_docosznak_value}">
                    </div>

                    <div>
                        <label><strong>СМСП:</strong></label>
                        <input type="text" id="smsp" value="{smsp_value}" size="15">
                    </div>
                </div>

                <div>
                    <label><strong>№ док.осн.зак.:</strong></label>
                    <input type="text" id="num_docosnzak" value="{num_docosnzak_value}" size="15">
                </div>

                <div>
                    <label><strong>Код основания некн.зак.:</strong></label>
                    <input type="text" id="OSTNEKONZAK" value="{OSTNEKONZAK_value}" size="15">
                </div>

                <div style="display: flex; gap: 50px;">
                    <div>
                        <label><strong>ОКПД2:</strong></label>
                        <input type="text" id="okpd2" value="{okpd2_value}" size="15">
                    </div>

                    <div>
                        <form action="#" style="margin-bottom: 0">
                            <label for="subectzak"><strong>SUBECTZAK</strong></label>
                            <select name="subectzak" id="subectzak">
                                <option value="">   </option>
                                <option value="001" {'selected' if subectzak_value == '001' else ''}>001</option>
                                <option value="002" {'selected' if subectzak_value == '002' else ''}>002</option>
                            </select>
                        </form> 
                    </div>
                </div>
            </div>
        </div>
        <!--КОНЕЦ РЕКВИЗИТОВ-------------->
        </div>       <!-- закрытие правого столбца-->

        <div>
            <button class="button-docs" onclick="showDocsDialog({contract_id})">Просмотр и сканирование документов</button>
        </div>

        </div>       <!--закрытие таблицы-->


        <!--ПРОСМОТР И СКАНИРОВНИЕ ДОКУМЕНТОВ-------------------->
        <dialog id="docsDialog" style="width: 80%; max-width: 1000px; padding: 10px; border: 2px solid #1073b7; border-radius: 8px;">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 5px;">
                <h2 style="color: #0952a0; margin-bottom: 5px;">Работа с отсканированными документами</h2>
                <button onclick="closeDocsDialog()" style="background: none; border: none; font-size: 24px; cursor: pointer;">✖</button>
            </div>

            <div style="background: #e4f0fb; padding: 3px; border-radius: 4px; margin-bottom: 10px;">
                <p><strong>ВНИМАНИЕ!</strong> Перед началом загрузки выберите в списке тип документации!</p>
            </div>

            <div style="margin-bottom: 20px;">
                <strong>Тип документации</strong>
                <div style="display: flex; flex-wrap: wrap; gap: 15px; margin-top: 10px;">
                    <label><input type="radio" name="docType" value="1"> Договор</label>
                    <label><input type="radio" name="docType" value="2"> Лист согласования</label>
                    <label><input type="radio" name="docType" value="3"> Дополнительные соглашения</label>
                    <label><input type="radio" name="docType" value="4"> Расчет арендной платы</label>
                    <label><input type="radio" name="docType" value="5"> Соглашение о расторжении</label>
                    <label><input type="radio" name="docType" value="6" checked> Платежки</label>
                    <label><input type="radio" name="docType" value="7"> Акт/накладная</label>
                </div>         
            </div>

            <h3>Загруженные файлы по этому договору</h3>

            <div style="max-height: 300px; overflow-y: auto; border: 1px solid #ddd; margin: 15px 0;">
                <table style="width: 100%; border-collapse: collapse;">
                    <thead style="background-color: #e4f0fb; position: sticky; top: 0;">
                        <tr>
                            <th style="padding: 8px; border: 1px solid #1073b7;">Тип док-ции</th>
                            <th style="padding: 8px; border: 1px solid #1073b7;">Файл</th>
                            <th style="padding: 8px; border: 1px solid #1073b7;">Размер (кб)</th>
                            <th style="padding: 8px; border: 1px solid #1073b7;">Дата созда.</th>
                            <th style="padding: 8px; border: 1px solid #1073b7;">Дата публ.</th>
                            <th style="padding: 8px; border: 1px solid #1073b7;">Статус</th>
                        </tr>
                    </thead>
                    <tbody id="filesTableBody">
                        <!-- Данные будут загружены через JS -->
                        <tr><td colspan="5" style="padding: 20px; text-align: center;">Загрузка файлов...</td></tr>
                    </tbody>
                </table>
            </div>

            <div style="display: flex; gap: 10px; justify-content: flex-end;">
                <button onclick="uploadDocument({contract_id})" class="button-save">Загрузить</button>
                <button onclick="openSelectedFile()" class="button-save">Открыть</button>
                <button onclick="deleteSelectedFile({contract_id})" class="button-save" style="background: #dc3545;">Удалить файл</button>
                <button onclick="publishSelectedFile()" class="button-save">Опубликовать/отменить</button>
            </div>
        </dialog>
        <!--ПРОСМОТР И СКАНИРОВНИЕ ДОКУМЕНТОВ-------------------->

        <div style="display: flex; justify-content: flex-start; align-items: flex-start; gap:20px;">        <!--открытие контейнера с нижней частью-->
        <!-- ТАБЛИЦА ОПЛАТ -->
        <div style="max-height: 200px; overflow-y: auto; width: 20%; border: 2px solid #1073b7; padding: 15px; border-radius: 8px; margin: 15px 0;">
            <div style="margin-bottom: 5px; display: flex; justify-content: space-between; align-items: center;">
                <h3 style="margin: 0px; color: #0952a0;">Оплаты по договору</h3>
                <div style="display: flex; gap: 15px; align-items: center;">
                    <button onclick="showAddPaymentForm()" class="button-adddeletepayment"><strong>+</strong></button>
                    <button onclick="deleteSelectedPayment()" class="button-adddeletepayment"><strong>-</strong></button>
                </div>
            </div>

            <div id="addPaymentForm" style="display: none; margin-bottom: 0px; padding: 10px; background: #f8f9fa; border-radius: 4px;">
                <input type="text" id="newPaymentSum" placeholder="Сумма" style="width: 45%; padding: 5px; margin-right: 5px;">
                <input type="date" id="newPaymentDate" style="width: 45%; padding: 5px; margin-right: 5px;">
                <div style="margin-top: 3px;"><button onclick="saveNewPayment({contract_id})" style="background: #1073b7; color: white; border: none; padding: 5px 10px; border-radius: 4px;">Сохранить</button>
                <button onclick="hideAddPaymentForm()" style="background: #6c757d; color: white; border: none; padding: 5px 10px; border-radius: 4px;">Отмена</button></div>
            </div>

            <table style="width: 100%; border-collapse: collapse;">
                <thead>
                    <tr style="background-color: #e4f0fb;">
                        <th style="padding: 8px; border: 1px solid #1073b7;">✓</th>
                        <th style="padding: 8px; border: 1px solid #1073b7;">Сумма</th>
                        <th style="padding: 8px; border: 1px solid #1073b7;">Дата</th>
                    </tr>
                </thead>
                <tbody>
        """
    if payments:
        for payment in payments:
            html += f"""
                    <tr>
                        <td style="padding: 5px; border: 1px solid #ddd; text-align: center;">
                            <input type="checkbox" class="payment-checkbox" value="{payment['id_opl']}">
                        </td>
                        <td style="padding: 5px; border: 1px solid #ddd;">{payment['s_opl']}</td>
                        <td style="padding: 5px; border: 1px solid #ddd;">{payment['d_opl'].strftime('%Y-%m-%d') if payment['d_opl'] else ''}</td>
                    </tr>
                """
    else:
        html += """
                    <tr>
                        <td colspan="3" style="padding: 20px; text-align: center; color: #666;">Оплаты по договору отсутствуют</td>
                    </tr>
            """
    html += """
                </tbody>
            </table>
        </div>       
        <!-- КОНЕЦ ТАБЛИЦЫ ОПЛАТ--------------------->

        <script>
        function showAddPaymentForm() {
            document.getElementById('addPaymentForm').style.display = 'block';
        }

        function hideAddPaymentForm() {
            document.getElementById('addPaymentForm').style.display = 'none';
            document.getElementById('newPaymentSum').value = '';
            document.getElementById('newPaymentDate').value = '';
        }

        function saveNewPayment(contractId) {
            const sum = document.getElementById('newPaymentSum').value;
            const date = document.getElementById('newPaymentDate').value;

            if (!sum || !date) {
                alert('Заполните сумму и дату');
                return;
            }

            fetch(`/api/contract/${contractId}/add_payment`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ summa: sum, date: date })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    location.reload();
                } else {
                    alert('Ошибка при добавлении оплаты');
                }
            });
        }

        function deleteSelectedPayment() {
            const checkboxes = document.querySelectorAll('.payment-checkbox:checked');
            const selectedIds = Array.from(checkboxes).map(cb => cb.value);

            if (selectedIds.length === 0) {
                alert('Выберите оплаты для удаления');
                return;
            }

            if (confirm(`Удалить ${selectedIds.length} запись(ей)?`)) {
                fetch('/api/contract/delete_payments', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ payment_ids: selectedIds })
                })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        location.reload();
                    } else {
                        alert('Ошибка при удалении');
                    }
                });
            }
        }

        // Показать диалог с документами
        function showDocsDialog(contractId) {
            const dialog = document.getElementById('docsDialog');
            dialog.showModal();
            loadContractFiles(contractId);
        }

        // Закрыть диалог
        function closeDocsDialog() {
            document.getElementById('docsDialog').close();
        }

        // Загрузить список файлов
        function loadContractFiles(contractId) {
            fetch(`/api/contract/${contractId}/files`)
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        renderFilesTable(data.files, contractId);
                    } else {
                        alert('Ошибка загрузки файлов: ' + data.error);
                    }
                });
        }

        // Отобразить таблицу файлов
        function renderFilesTable(files, contractId) {
            const tbody = document.getElementById('filesTableBody');

            if (files.length === 0) {
                tbody.innerHTML = '\\ <td colspan="6" style="padding: 20px; text-align: center;">Нет загруженных файлов<\/td><\/tr>';
                return;
            }

            let html = '';
            files.forEach(file => {
                const isPublished = file.published === '1';

                html += `
                    <tr class="file-row" data-file-path="${file.full_path}" style="cursor: pointer;">
                        <td style="padding: 5px; border: 1px solid #ddd;">${file.doc_type || ''}<\/td>
                        <td style="padding: 5px; border: 1px solid #ddd;">${file.filename}<\/td>
                        <td style="padding: 5px; border: 1px solid #ddd;">${file.size_kb}<\/td>
                        <td style="padding: 5px; border: 1px solid #ddd;">${file.created || ''}<\/td>
                        <td style="padding: 5px; border: 1px solid #ddd;">${isPublished ? file.published_date : '—'}<\/td>
                        <td style="padding: 5px; border: 1px solid #ddd;">
                            <button onclick="togglePublish(${contractId}, '${file.full_path}', ${!isPublished})" 
                                    class="button-save" 
                                    style="background: ${isPublished ? '#dc3545' : '#28a745'}; padding: 3px 10px; font-size: 12px;">
                                ${isPublished ? 'Отменить публикацию' : 'Опубликовать'}
                            </button>
                         <\/td>
                    <\/tr>
                `;
            });
            tbody.innerHTML = html;

            // ========== ДОБАВЛЯЕМ ОБРАБОТЧИКИ КЛИКА НА СТРОКИ ==========
            const rows = document.querySelectorAll('.file-row');
            rows.forEach(row => {
                row.addEventListener('click', function(e) {
                    //Если кликнули не по кнопке, выделяем строку
                    if (e.target.tagName !== 'BUTTON') {
                        selectFileRow(this);
                    }
                });
            });
        }

        // Функция выделения строки
        function selectFileRow(rowElement) {
            // Снимаем выделение со всех строк
            const allRows = document.querySelectorAll('.file-row');
            allRows.forEach(row => {
                row.style.backgroundColor = '';
            });

            // Выделяем выбранную строку
            rowElement.style.backgroundColor = '#e4f0fb';

            // Сохраняем путь к файлу
            selectedFilePath = rowElement.getAttribute('data-file-path');
        }

        function togglePublish(contractId, filePath, publish) {
            if (publish) {
                // Опубликовать - запрашиваем дату
                const today = new Date().toISOString().slice(0, 10);
                const date = prompt('Введите дату публикации (ГГГГ-ММ-ДД):', today);
                if (!date) return;

                // Проверяем формат даты
                if (!/^\d{4}-\d{2}-\d{2}$/.test(date)) {
                    alert('Неверный формат даты. Используйте ГГГГ-ММ-ДД');
                    return;
                }

                // Преобразуем в формат yyyymmdd
                const datePubl = date.replace(/-/g, '');

                fetch(`/api/contract/${contractId}/file/publish`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ 
                        file_path: filePath, 
                        publish: true, 
                        date_publ: datePubl 
                    })
                })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        // Обновляем список файлов
                        loadContractFiles(contractId);
                    } else {
                        alert('Ошибка: ' + data.message);
                    }
                })
                .catch(error => {
                    alert('Ошибка: ' + error);
                });
            } else {
                // Отменить публикацию - спрашиваем подтверждение
                if (confirm('Отменить публикацию файла?')) {
                    fetch(`/api/contract/${contractId}/file/publish`, {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({ 
                            file_path: filePath, 
                            publish: false 
                        })
                    })
                    .then(response => response.json())
                    .then(data => {
                        if (data.success) {
                            loadContractFiles(contractId);
                        } else {
                            alert('Ошибка: ' + data.message);
                        }
                    })
                    .catch(error => {
                        alert('Ошибка: ' + error);
                    });
                }
            }
        }

        let selectedFilePath = null;  

        function selectFileRow(rowElement) {
            // Снимаем выделение со всех строк
            const allRows = document.querySelectorAll('.file-row');
            allRows.forEach(row => {
                row.style.backgroundColor = '';
            });

            // Выделяем выбранную строку
            rowElement.style.backgroundColor = '#e4f0fb';

            // Сохраняем путь к файлу
            selectedFilePath = rowElement.getAttribute('data-file-path');
        }

        function openSelectedFile() {
            if (!selectedFilePath) {
                alert('Выберите файл');
                return;
            }
            window.open(`/api/contract/open_file?path=${encodeURIComponent(selectedFilePath)}`, '_blank');
        }

        function deleteSelectedFile(contractId) {
            if (!selectedFilePath) {
                alert('Выберите файл');
                return;
            }
            if (!confirm('Удалить файл?')) return;

            fetch(`/api/contract/${contractId}/files/delete`, {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ file_path: selectedFilePath })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    loadContractFiles(contractId);
                    selectedFilePath = null;
                } else {
                    alert('Ошибка удаления: ' + data.error);
                }
            });
        }

        function uploadDocument(contractId) {
        const selectedDocType = document.querySelector('input[name="docType"]:checked');
        if (!selectedDocType) {
            alert('Выберите тип документации');
            return;
        }

        const docType = selectedDocType.value;
        console.log('Выбран тип:', docType);

        const input = document.createElement('input');
        input.type = 'file';
        input.accept = '.pdf,.jpg,.png,.doc,.docx';

        input.onchange = async (event) => {
            const file = event.target.files[0];
            if (!file) return;

            console.log('Выбран файл:', file.name, 'Размер:', file.size);

            const formData = new FormData();
            formData.append('doc_type', docType);
            formData.append('file', file);

            try {
                console.log('Отправка запроса...');
                const response = await fetch(`/api/contract/${contractId}/upload`, {
                    method: 'POST',
                    body: formData
                });

                console.log('Статус ответа:', response.status);
                const data = await response.json();
                console.log('Ответ сервера:', data);

                if (data.success) {
                    alert('Файл успешно загружен');
                    loadContractFiles(contractId);
                } else {
                    alert('Ошибка загрузки: ' + data.message);
                }
            } catch (error) {
                console.error('Ошибка:', error);
                alert('Ошибка: ' + error);
            }
        };
    input.click();
}
        </script>
        """

    html += f"""
        <!-- ПОЛЯ МЕЖДУ ТАБЛИЦАМИ---------------------->
        <div style="line-height: 23px; margin-top: 25px;">
            <div style="display: flex; gap: 5px; align-items: center; margin-top: 0px;">
                <input type="checkbox" id="smsp_okz" {smsp_okz_checked}>
                <label for="smsp_okz"><strong>Загрузка среди СМСП</strong></label>
            </div>

            <div style="display: flex; gap: 5px; justify-content: space-between; margin-bottom: 5px;">
                <label for="summaokz"><strong>Сумма договора ОКЗ:</strong></label>
                <input type="text" id="summaokz" value="{s_dog_okz_value}" size="15">
            </div>

            <div style="display: flex; gap: 5px; justify-content: space-between; margin-bottom: 5px;">
                <label for="summasds"><strong>Сумма c ДС:</strong></label>
                <input type="text" id="summasds" value="{s_ds_value}" size="15">
            </div>

            <div style="display: flex; gap: 5px; justify-content: space-between; margin-bottom: 5px;">
                <label for="dateizv"><strong>Дата извещения:</strong></label>
                <input type="date" id="dateizv" value="{date_izv_value}" size="15">
            </div>

            <div style="display: flex; gap: 5px; align-items: center; margin-bottom: 5px;">
                <input type="checkbox" id="agent" {agent_checked}>
                <label for="agent"><strong>Агентский договор</strong></label>
            </div>
            <div style="margin-bottom: 5px;">
                <strong>Системный номер:</strong> {contract.get('Системный номер', 'Нет данных')}
            </div>
            <div style="display: flex; gap: 5px; align-items: center; margin-bottom: 5px;">
                <label for="d_work"><strong>Рабочая дата:</strong></label>
                <input type="date" id="d_work" value="{d_work_value}">    
            </div>

            <div style="display: flex; gap: 5px; align-items: center; margin-bottom: 5px;">
                <label for="predlog_txt"><strong>№ предложения:</strong></label>
                <input type="text" id="predlog_txt" value="{predlog_txt_value}">    
            </div>
        </div>
        <!-- ПОЛЯ МЕЖДУ ТАБЛИЦАМИ---------------------->
        """
    html += f"""
        <!-- ТАБЛИЦА ИЗ DSPROC -->
        <div style="max-height: 200px; overflow-y: auto; width: 30%; border: 2px solid #1073b7; padding: 15px; border-radius: 8px; margin: 15px 0;">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">
                <h3 style="margin: 0; color: #0952a0;">Дополнительные соглашения</h3>
            </div>

            <table style="width: 100%; border-collapse: collapse;">
                <thead>
                    <tr style="background-color: #e4f0fb;">
                        <th style="padding: 8px; border: 1px solid #1073b7;">№ дс</th>
                        <th style="padding: 8px; border: 1px solid #1073b7;">Дата рег.</th>
                        <th style="padding: 8px; border: 1px solid #1073b7;">Дата нач.</th>
                        <th style="padding: 8px; border: 1px solid #1073b7;">Дата оконч.</th>
                        <th style="padding: 8px; border: 1px solid #1073b7;">azec_ds</th>
                    </tr>
                </thead>
                <tbody>
        """
    if ds_data:
        for item in ds_data:
            html += f"""
                    <tr onclick="showPaymentDialog('{item.get('numds', '')}', '{item.get('drds', '')}', '{item.get('dnds', '')}', '{item.get('dkds', '')}', '{item.get('azes_ds', '')}', '{item.get('viddopsog', '')}', '{item.get('sod', '')}',{contract_id})">
                        <td style="padding: 5px; border: 1px solid #ddd;">{item.get('numds', '')}</td>
                        <td style="padding: 5px; border: 1px solid #ddd;">{item.get('drds', '')}</td>
                        <td style="padding: 5px; border: 1px solid #ddd;">{item.get('dnds', '')}</td>
                        <td style="padding: 5px; border: 1px solid #ddd;">{item.get('dkds', '')}</td>
                        <td style="padding: 5px; border: 1px solid #ddd;">{item.get('azes_ds', '')}</td>
                    </tr>
                """
    else:
        html += """
                    <tr>
                        <td colspan="5" style="padding: 20px; text-align: center; color: #666;">Данные отсутствуют</td>
                    </tr>
            """
    html += """
                </tbody>
            </table>
        </div>
        <!-- КОНЕЦ ТАБЛИЦЫ ИЗ DSPROC -->

        <!--ДИАЛОГОВОЕ ОКНО ПРИ НАЖАТИИ НА ОПЛАТУ В ТАБЛИЦЕ ДОПОЛНИТЕЛЬНЫЕ СОГЛАШЕНИЯ------------------->
        <dialog id="paymentDialog" style="width: 20%; padding: 10px; border: 2px solid #1073b7; border-radius: 8px;">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 5px;">
                <h3 style="color: #0952a0; margin-bottom: 5px;">Дополнительные соглашения</h3>
                <button onclick="closePaymentDialog()" style="background: none; border: none; font-size: 24px; cursor: pointer;">✖</button>
            </div>

            <div id="paymentDialogContent"></div>

            <div style="margin-top: 15px; text-align: right;">
                <button onclick="saveAzecDs()" class="button-save" style="background: #28a745;">ОК</button>
            </div>
        </dialog>
        <!--КОНЕЦ ДИАЛОГОВОГО ОКНА ПРИ НАЖАТИИ НА ОПЛАТУ В ТАБЛИЦЕ ДОПОЛНИТЕЛЬНЫЕ СОГЛАШЕНИЯ------------------->

        <script>
        function showPaymentDialog(numds, drds, dnds, dkds, azes_ds, viddopsog, sod, contractId) {
            const dialog = document.getElementById('paymentDialog');
            dialog.setAttribute('data-numds', numds);
            dialog.setAttribute('data-contract-id', contractId);

            // Формируем HTML с данными
            const content = `
                <div><strong>№ ДС:</strong> ${numds}</div>
                <div><strong>Дата регистрации:</strong> ${drds}</div>
                <div><strong>Дата начала:</strong> ${dnds}</div>
                <div><strong>Дата окончания:</strong> ${dkds}</div>
                <div style="display: flex; gap: 5px; align-items: center;">
                    <strong>AZES_DS:</strong>
                    <input type="text" id="edit_azes_ds" size="20" value="${azes_ds}" style="height: 20px; padding: 5px;" >
                </div>
                <div><strong>VIDDOPSOG:</strong> ${viddopsog}</div>
                <div><strong>Содержание:</strong> ${sod}</div>
            `;

            // Вставляем в диалог
            document.getElementById('paymentDialogContent').innerHTML = content;

            // Открываем диалог
            document.getElementById('paymentDialog').showModal();
        }

        function closePaymentDialog() {
            document.getElementById('paymentDialog').close();
        }

        function saveAzecDs() {
            const dialog = document.getElementById('paymentDialog');
            const numds = dialog.getAttribute('data-numds');
            const contractId = dialog.getAttribute('data-contract-id');
            const newValue = document.getElementById('edit_azes_ds').value;

            if (!numds || !contractId) {
                alert('Ошибка: не удалось определить номер ДС или ID договора');
                return;
            }

            fetch('/api/ds/update_azec', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ 
                    numds: numds, 
                    azes_ds: newValue,
                    contract_id: contractId
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    alert('Сохранено успешно!');
                    closePaymentDialog();
                    // Обновляем таблицу на странице
                    location.reload();
                } else {
                    alert('Ошибка: ' + data.message);
                }
            })
            .catch(error => {
                alert('Ошибка: ' + error);
            });
        }
        </script>
        """

    html += f"""

        <!-- ДВЕ КНОПКИ -->
        <div style="display: flex; width: 97%; justify-content: space-between; position: fixed; align-items: center; bottom: 20px;">
            <div>
                <a href="{back_url}" class="button-back-page">Назад к списку</a>
            </div> 
            <div>   
                <button id="saveButton" onclick="savepar()" class="button-save">Cохранить</button>
            </div>
        </div>


    <script>
    function savepar() {{
        const saveButton = document.getElementById('saveButton')    //ищется кнопка сохранить 
        saveButton.textContent = "Сохранение...";   //смена текста после нажатия кнопки 
        saveButton.disabled = true;    //не дает второй раз надать пока не загорится снова синим 

        const data = {{      //сбор всех полей которые изменили и не изменили для дальнейшего сохранения 
            konk: document.getElementById('konkCheckbox').checked ? 1 : 0,
            prol: document.getElementById('prol').checked ? 1 : 0,      
            beznds: document.getElementById('beznds').checked ? 1 : 0,  
            opl: document.getElementById('opl').checked ? 1 : 0,        
            eis: document.getElementById('eis').checked ? 1 : 0,      
            statusD: document.getElementById('status').value, 
            d_end: document.getElementById('d_end').value,

            sposobzak: document.getElementById('sposobzak').value,
            VIdZAK: document.getElementById('VIdZAK').value,
            numzak: document.getElementById('numzak').value,
            predlog: document.getElementById('predlog').value === 'opened' ? 1 : 0,
            dat_docosznak: document.getElementById('dat_docosznak').value,
            num_docosnzak: document.getElementById('num_docosnzak').value,
            smsp: document.getElementById('smsp').value,
            OSTNEKONZAK: document.getElementById('OSTNEKONZAK').value,
            okpd2: document.getElementById('okpd2').value,
            subectzak: document.getElementById('subectzak').value,

            num_z: document.getElementById('num_z').value,
            num_z_el: document.getElementById('numzalel').value,
            pr_z: document.getElementById('przakcheckbox').checked ? 1 : 0,
            pr_z_osn: document.getElementById('osn').value, 
            gpz: document.getElementById('gpz').value,
            uid: document.getElementById('uid').value,
            ppz: document.getElementById('ppz').value,    

            s_dog_okz: document.getElementById('summaokz').value, 
            s_ds: document.getElementById('summasds').value,
            date_izv: document.getElementById('dateizv').value, 

            agent: document.getElementById('agent').checked ? 1 : 0,
            smsp_okz: document.getElementById('smsp_okz').checked ? 1 : 0,
            d_work: document.getElementById('d_work').value,
            predlog_txt: document.getElementById('predlog_txt').value
            publ: document.getElementById('publ_checkbox').checked ? 1 : 0,
            publ_d: document.getElementById('publ_date').value
        }};
        console.log('Данные для отправки:', data);

        //отправляем запрос на сервер, method: 'POST' - отправить данные
        fetch('/api/contract/{contract_id}/update', {{      //fetch - функция для отправление запросов на сервер, в нашем случае на адрес которые делает сохранение полей 
            method: 'POST',       //метод post - отправление/обновление данных. есть ещё get - он для получения данных
            headers: {{ 'Content-Type': 'application/json' }},   //пищет что мы отправляем. в нашем случае json строку
            body: JSON.stringify(data)     //изменяет js в json
        }})

        .then(response => response.json())    //меняем json строку в js объект 
        .then(data => {{
            if (data.success) {{     //если успешно поменялось поле в таблице изменяем текст и цвет кнопки
                saveButton.textContent = "Сохранено";
                saveButton.style.background = "#28a745";
            }} else {{
                saveButton.textContent = "Ошибка"; 
                saveButton.style.background = "#dc3545";
            }}

            setTimeout(() => {{     //функция позволяет запустить код с задержкой, в нашем случае изменить текст и цвет кнопки 
                saveButton.textContent = "Сохранить";
                saveButton.style.background = "#1073b7";
                saveButton.disabled = false;        //чтобы второй раз не нажал на сохранение иначе второй раз запрос будет проходить 
            }}, 4000);
        }});
    }}
    </script>


        <!-- ТАБЛИЦА ОПЛАТ ИЗ 1С -->
        <div style="max-height: 200px; overflow-y: auto; width: 20%; padding: 15px; margin: 15px 0; border: 2px solid #1073b7; border-radius: 8px;">
            <h3 style="margin: 0 0 10px 0; color: #0952a0;">Оплаты из 1С на сумму {total_sum_1c:.2f}</h3>

            <table style="width: 100%; border-collapse: collapse;">
                <thead>
                    <tr style="background-color: #e4f0fb;">
                        <th style="padding: 8px; border: 1px solid #1073b7;">Сумма</th>
                        <th style="padding: 8px; border: 1px solid #1073b7;">Дата</th>
                    </tr>
                </thead>
                <tbody>
        """
    if payments1C:
        for payment in payments1C:
            html += f"""
                    <tr>
                        <td style="padding: 5px; border: 1px solid #ddd;">{payment.get('s_opl', '')}</td>
                        <td style="padding: 5px; border: 1px solid #ddd;">{payment['d_opl'].strftime('%Y-%m-%d') if payment['d_opl'] else ''}</td>
                    </tr>
                """
    else:
        html += """
                    <tr>
                        <td colspan="2" style="padding: 20px; text-align: center; color: #666;">Оплаты из 1С отсутствуют</td>
                    </tr>
                </tbody>
            </table>
        </div>
        <!-- КОНЕЦ ТАБЛИЦЫ ОПЛАТ ИЗ 1С -->
        </div>           <!--закрытие контейнера с нижней частью-->
    </div>             <!-- закрытие общего контейнера с фоном -->

    </body>
    </html>
    """
    return html


@app.post("/api/contract/{contract_id}/update")
async def update_contract_api(contract_id: int, request: Request):
    # собирает параметры перед нажатием кнопки сохранить
    try:
        data = await request.json()

        konk = data.get('konk', 0)
        prol = data.get('prol', 0)
        beznds = data.get('beznds', 0)
        opl = data.get('opl', 0)
        eis = data.get('eis', 0)
        statusD = data.get('statusD', 1)
        d_end = data.get('d_end', None)

        num_z = data.get('num_z', '')
        num_z_el = data.get('num_z_el', '')
        pr_z = data.get('pr_z', 0)
        pr_z_osn = data.get('pr_z_osn', '')
        gpz = data.get('gpz', '')
        uid = data.get('uid', '')
        ppz = data.get('ppz', '')

        sposobzak = data.get('sposobzak', '')
        VIdZAK = data.get('VIdZAK', '')
        numzak = data.get('numzak', '')
        predlog = data.get('predlog', 0)
        dat_docosznak = data.get('dat_docosznak', '')
        smsp = data.get('smsp', '')
        num_docosnzak = data.get('num_docosnzak', '')
        OSTNEKONZAK = data.get('OSTNEKONZAK', '')
        okpd2 = data.get('okpd2', '')
        subectzak = data.get('subectzak', '')

        s_dog_okz = data.get('s_dog_okz', '')
        s_ds = data.get('s_ds', '')
        date_izv = data.get('date_izv', '')

        agent = data.get('agent', 0)
        smsp_okz = data.get('smsp_okz', 0)
        d_work = data.get('d_work', '')
        predlog_txt = data.get('predlog_txt', '')
        publ = data.get('publ', 0)
        publ_d = data.get('publ_d', '')

        success = update_par(
            request,
            contract_id=contract_id,

            konk=konk,
            prol=prol,
            beznds=beznds,
            opl=opl,
            eis=eis,
            statusD=statusD,
            d_end=d_end,

            sposobzak=sposobzak,
            VIdZAK=VIdZAK,
            numzak=numzak,
            predlog=predlog,
            dat_docosznak=dat_docosznak,
            num_docosnzak=num_docosnzak,
            smsp=smsp,
            OSTNEKONZAK=OSTNEKONZAK,
            okpd2=okpd2,
            subectzak=subectzak,

            num_z=num_z,
            num_z_el=num_z_el,
            pr_z=pr_z,
            pr_z_osn=pr_z_osn,
            gpz=gpz,
            uid=uid,
            ppz=ppz,

            s_dog_okz=s_dog_okz,
            s_ds=s_ds,
            date_izv=date_izv,

            agent=agent,
            smsp_okz=smsp_okz,
            d_work=d_work,
            predlog_txt=predlog_txt,

            publ=publ,
            publ_d=publ_d
        )

        if success:
            return {"success": True, "message": "Статус сохранен"}
        else:
            return {"success": False, "message": "Ошибка сохранения"}

    except Exception as e:
        print(f"Error in update_contract_api: {e}")
        return {"success": False, "message": f"Ошибка: {str(e)}"}


# для добавления оплаты в первую таблицу
@app.post("/api/contract/{contract_id}/add_payment")
async def add_payment(contract_id: int, request: Request):
    try:
        data = await request.json()
        summa = data.get('summa')
        date = data.get('date')

        success = add_dog_payment(request, contract_id, summa, date)

        return {"success": success}
    except Exception as e:
        print(f"Error adding payment: {e}")
        return {"success": False, "message": str(e)}


# для удаления оплаты из первой таблицы
@app.post("/api/contract/delete_payments")
async def delete_payments(request: Request):
    try:
        data = await request.json()
        payment_ids = data.get('payment_ids', [])

        success = delete_dog_payments(request, payment_ids)

        return {"success": success}
    except Exception as e:
        print(f"Error deleting payments: {e}")
        return {"success": False, "message": str(e)}


@app.get("/api/contract/{contract_id}/files")
async def get_contract_files_api(contract_id: int, request: Request):
    try:
        contract = str(contract_id)
        if not contract:
            return {"success": False, "error": "Договор не найден"}

        contract_num = str(contract_id)
        files = get_contract_files(request, contract_id)

        return {"success": True, "files": files}
    except Exception as e:
        return {"success": False, "error": str(e)}


# удаление файла пдф договора
@app.post("/api/contract/{contract_id}/files/delete")
async def delete_contract_file(request: Request):
    try:
        data = await request.json()
        file_path = data.get('file_path')

        if not file_path or not os.path.exists(file_path):
            return {"success": False, "error": "Файл не найден"}

        os.remove(file_path)
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}


# открыть и скачать файл пдф договора
@app.get("/api/contract/open_file")
async def open_file(request: Request, path: str):
    try:
        if not os.path.exists(path):
            return {"success": False, "error": "Файл не найден"}

        # Возвращаем файл с заголовком для скачивания
        return FileResponse(
            path=path,
            filename=os.path.basename(path),
            media_type='application/octet-stream',  # <- важно!
            headers={
                "Content-Disposition": f"attachment; filename={os.path.basename(path)}"
            }
        )
    except Exception as e:
        return {"success": False, "error": str(e)}


# загрузить файл в таблицу для договора
@app.post("/api/contract/{contract_id}/upload")
async def upload_file(
        contract_id: int,
        doc_type: str = Form(...),
        file: UploadFile = File(...)
):
    """Загрузить файл для договора"""
    try:
        base_path = '/mnt/oblgaz/system/contracts_archive/'
        contract_folder = os.path.join(base_path, str(contract_id))
        os.makedirs(contract_folder, exist_ok=True)

        # Считаем количество существующих файлов этого типа
        existing_files = [f for f in os.listdir(contract_folder)
                          if f.startswith(f"{contract_id}_{doc_type}_")]
        counter = len(existing_files) + 1

        ext = os.path.splitext(file.filename)[1]
        filename = f"{contract_id}_{doc_type}_{counter}{ext}"
        file_path = os.path.join(contract_folder, filename)

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        return {"success": True, "message": "Файл загружен", "filename": filename}

    except Exception as e:
        return {"success": False, "message": str(e)}

# публикация и отмена публикации
@app.post("/api/contract/{contract_id}/file/publish")
async def publish_file(contract_id: int, request: Request):
    try:
        data = await request.json()
        file_path = data.get('file_path')
        publish = data.get('publish', True)
        date_publ = data.get('date_publ', datetime.now().strftime('%Y%m%d'))

        # Получаем подключение к БД
        connection = get_user_connection(request)
        if not connection:
            return {"success": False, "message": "Нет подключения к базе данных"}

        cursor = connection.cursor()

        # парсим имя файла: 167536_6_1.pdf
        filename = os.path.basename(file_path)
        parts = filename.split('_')
        if len(parts) >= 3:
            type_file = parts[1]  # тип документации
            num_file = parts[2].split('.')[0]  # уникальный номер
        else:
            cursor.close()
            connection.close()
            return {"success": False, "message": "Неверный формат имени файла"}

        if publish:
            # Опубликовать
            query = """
                INSERT INTO dog_file_p (id_file, type_file, num_file, date_publ)
                VALUES (?, ?, ?, ?)
            """
            cursor.execute(query, contract_id, type_file, num_file, date_publ)
        else:
            # Отменить публикацию (удалить)
            query = """
                DELETE FROM dog_file_p 
                WHERE id_file = ? AND type_file = ? AND num_file = ?
            """
            cursor.execute(query, contract_id, type_file, num_file)

        connection.commit()
        cursor.close()
        connection.close()

        return {"success": True, "message": "Готово"}

    except Exception as e:
        print(f"Error in publish_file: {e}")
        return {"success": False, "message": str(e)}

# экспорт договоров в ексель
@app.get("/api/contract/export")
async def export_contracts(
        request: Request,
        numberdog: str = "",
        numberkontr: str = "",
        date_from: str = "",
        date_to: str = "",
        publ: str = "",
        sum_from: str = "",
        sum_to: str = "",
        podr: str = "",
        pr_dog: str = "",
        gazsrv: str = "",
        search_archive: str = ""
):
    try:
        # Получаем результаты поиска со всеми параметрами
        results = search_dog(
            request,
            numberdog,
            numberkontr,
            date_from,
            date_to,
            publ,
            sum_from,
            sum_to,
            podr,
            pr_dog,
            gazsrv,
            search_archive
        )

        if not results:
            return {"success": False, "message": "Нет данных для экспорта"}

        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реестр договоров"  # короткое фиксированное название

        headers = ['ID договора', 'Номер договора', '№ контрагента', 'Дата договора', 'Контрагент', 'Предмет договора']
        header_font = Font(bold=True, color="000000")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        for row, contract in enumerate(results, 2):
            # ID договора (всегда должен быть)
            ws.cell(row=row, column=1, value=contract.get('ID договора', ''))
            ws.cell(row=row, column=2, value=contract.get('Номер договора', ''))
            ws.cell(row=row, column=3, value=contract.get('№ контрагента', ''))
            ws.cell(row=row, column=4, value=contract.get('Дата договора', ''))
            ws.cell(row=row, column=5, value=contract.get('Контрагент', ''))

            # Для предмета договора - с обработкой ошибки
            try:
                ws.cell(row=row, column=6, value=contract.get('Предмет договора', ''))
            except Exception as e:
                print(f"Ошибка в строке {row}: {e}")
                ws.cell(row=row, column=6, value="[ошибка отображения]")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].auto_size = True

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=contracts_report.xlsx"}
        )
    except Exception as e:
        print(f"Error in export: {e}")
        return {"success": False, "message": str(e)}

@app.post("/api/ds/update_azec")
async def update_azec_ds(request: Request):
    try:
        data = await request.json()
        numds = data.get('numds')
        azes_ds = data.get('azes_ds')
        contract_id = data.get('contract_id')

        if not numds or not contract_id:
            return {"success": False, "message": "Не указан номер ДС или ID договора"}

        connection = get_user_connection(request)
        if not connection:
            return {"success": False, "message": "Нет подключения к БД"}
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE ds 
            SET azes_ds = ? 
            WHERE numds = ? AND iddog = ?
        """, azes_ds, numds, contract_id)
        connection.commit()
        cursor.close()
        connection.close()
        return {"success": True, "message": "Данные обновлены"}

    except Exception as e:
        (f"Error in save_ds: {e}")
        return {"success": False, "message": str(e)}

#страница выбора программы
@app.get("/choose", response_class=HTMLResponse)
def choose_page(request: Request):
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=303)

    username = user.get("login", "Пользователь")

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Выбор программы</title>
        <style>
            .choose_body {{
                font-family: Arial, sans-serif;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                margin: 0;
                background: #f0f2f5;
            }}
            .choose_container {{
                text-align: center;
                background: white;
                padding: 40px;
                border-radius: 10px;
                box-shadow: 0 0 20px rgba(0,0,0,0.1);
            }}
            .choose_h1 {{
                color: #0952a0;
                margin-bottom: 10px;
            }}
            .welcome {{
                color: #666;
                margin-bottom: 30px;
                font-size: 16px;
            }}
            .buttons {{
                display: flex;
                gap: 30px;
                justify-content: center;
            }}
            .btn {{
                background: #1073b7;
                color: white;
                border: none;
                padding: 15px 30px;
                font-size: 18px;
                border-radius: 5px;
                cursor: pointer;
                text-decoration: none;
                display: inline-block;
            }}
            .btn:hover {{
                background: #0952a0;
            }}
            .btn-secondary {{
                background: #6c757d;
            }}
            .btn-secondary:hover {{
                background: #5a6268;
            }}
            .logout {{
                margin-top: 30px;
            }}
            .logout a {{
                color: #999;
                text-decoration: none;
                font-size: 14px;
            }}
            .logout a:hover {{
                color: #666;
            }}
        </style>
    </head>
    <body class="choose_body">
        <div class="choose_container">
            <h1 style="choose_h1">Выберите программу</h1>
            <div class="welcome">Добро пожаловать, <strong>{username}</strong>!</div>
            <div class="buttons">
                <a href="/" class="btn">Расходные договоры и закупки</a>
                <button onclick="showMessage()" class="btn btn-secondary">Другая программа</button>
            </div>
            <div class="logout">
                <a href="/logout">Выйти</a>
            </div>
        </div>

        <script>
            function showMessage() {{
                alert('Программа в разработке. Скоро будет доступна.');
            }}
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

if __name__ == '__main__':
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8001,
        ssl_keyfile="key.pem",
        ssl_certfile="cert.pem"
    )  # ip aдрес моего компьютера 192.168.1.228, по умолчанию 127.0.0.1, https://192.168.1.228:8001